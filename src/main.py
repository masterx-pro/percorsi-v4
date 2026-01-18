"""
Percorsi Android Pro v3.2 - Ottimizzatore di Percorsi
Autore: Mattia Prosperi

- Gestione Excel/CSV robusta
- Fallback CSV se openpyxl non disponibile
- File picker Android
"""

import os
import json
import threading
import traceback
from math import radians, sin, cos, atan2, sqrt, pi, log, tan
from datetime import datetime

from kivy.app import App
from kivy.lang import Builder
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.utils import platform
from kivy.properties import StringProperty, ListProperty, BooleanProperty
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.progressbar import ProgressBar
from kivy.uix.stencilview import StencilView
from kivy.graphics import Color, Line, Ellipse, Rectangle
from kivy.logger import Logger

# ==================== IMPORTS CONDIZIONALI ====================
try:
    import requests
    REQUESTS_AVAILABLE = True
except:
    REQUESTS_AVAILABLE = False

# Openpyxl - prova import con gestione errori dettagliata
OPENPYXL_AVAILABLE = False
OPENPYXL_ERROR = ""
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
    Logger.info("Percorsi: openpyxl DISPONIBILE")
except ImportError as e:
    OPENPYXL_ERROR = str(e)
    Logger.warning(f"Percorsi: openpyxl NON disponibile: {e}")
except Exception as e:
    OPENPYXL_ERROR = str(e)
    Logger.error(f"Percorsi: Errore openpyxl: {e}")

try:
    from plyer import filechooser as plyer_fc
    PLYER_AVAILABLE = True
except:
    PLYER_AVAILABLE = False

# Cache distanze
DIST_CACHE = {}

def log_debug(msg):
    Logger.info(f"Percorsi: {msg}")

def log_error(msg):
    Logger.error(f"Percorsi: {msg}")

# ==================== MATH FUNCTIONS ====================
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = radians(lat1), radians(lat2)
    dp = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dp/2)**2 + cos(p1)*cos(p2)*sin(dl/2)**2
    return R * 2 * atan2(sqrt(a), sqrt(1-a))

def get_distance(lat1, lon1, lat2, lon2, mode="haversine"):
    key = (round(lat1,5), round(lon1,5), round(lat2,5), round(lon2,5), mode)
    if key in DIST_CACHE:
        return DIST_CACHE[key]
    
    if mode == "osrm" and REQUESTS_AVAILABLE:
        try:
            url = f"https://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}?overview=false"
            r = requests.get(url, timeout=10)
            d = r.json()
            if d.get("code") == "Ok":
                dist = int(d["routes"][0]["distance"])
                DIST_CACHE[key] = dist
                return dist
        except:
            pass
    
    dist = int(haversine(lat1, lon1, lat2, lon2))
    DIST_CACHE[key] = dist
    return dist

def build_matrix(coords, mode="haversine", cb=None):
    n = len(coords)
    matrix = {i: {j: 0 for j in range(n)} for i in range(n)}
    total = n * (n-1) // 2
    done = 0
    for i in range(n):
        for j in range(i+1, n):
            d = get_distance(coords[i][0], coords[i][1], coords[j][0], coords[j][1], mode)
            matrix[i][j] = d
            matrix[j][i] = d
            done += 1
            if cb and total > 0:
                cb(done / total)
    return matrix

def nearest_neighbor(matrix, start=0):
    n = len(matrix)
    if n <= 1:
        return list(range(n))
    visited = [False] * n
    tour = [start]
    visited[start] = True
    for _ in range(n-1):
        curr = tour[-1]
        nearest, nd = None, float('inf')
        for j in range(n):
            if not visited[j] and matrix[curr][j] < nd:
                nearest, nd = j, matrix[curr][j]
        if nearest is not None:
            tour.append(nearest)
            visited[nearest] = True
    return tour

def two_opt(tour, matrix, max_iter=500):
    n = len(tour)
    if n < 4:
        return tour
    improved = True
    iters = 0
    while improved and iters < max_iter:
        improved = False
        iters += 1
        for i in range(n-2):
            for j in range(i+2, n):
                if j == n-1 and i == 0:
                    continue
                a, b, c, d = tour[i], tour[i+1], tour[j], tour[(j+1) % n]
                if matrix[a][b] + matrix[c][d] > matrix[a][c] + matrix[b][d]:
                    tour[i+1:j+1] = reversed(tour[i+1:j+1])
                    improved = True
    return tour

def solve_tsp(coords, mode="haversine", cb=None):
    if len(coords) < 2:
        return list(range(len(coords))), 0, [0]*len(coords)
    if cb: cb(0.1)
    matrix = build_matrix(coords, mode, lambda p: cb(0.1 + p*0.5) if cb else None)
    if cb: cb(0.7)
    tour = nearest_neighbor(matrix, 0)
    tour = two_opt(tour, matrix)
    if cb: cb(0.95)
    distances = [0]
    for i in range(1, len(tour)):
        distances.append(matrix[tour[i-1]][tour[i]])
    if cb: cb(1.0)
    return tour, sum(distances), distances

def divide_for_operators(indices, coords, num_ops, items_per, mode="haversine"):
    if num_ops <= 1 or len(indices) == 0:
        return [indices]
    
    centroid = (sum(coords[i][0] for i in indices)/len(indices),
                sum(coords[i][1] for i in indices)/len(indices))
    
    sorted_idx = sorted(indices, key=lambda i: haversine(coords[i][0], coords[i][1], centroid[0], centroid[1]))
    
    assignments = [[] for _ in range(num_ops)]
    for i, idx in enumerate(sorted_idx):
        op = i % num_ops
        if len(assignments[op]) < items_per:
            assignments[op].append(idx)
    
    result = []
    for op_idx in assignments:
        if len(op_idx) > 1:
            op_coords = [coords[i] for i in op_idx]
            tour, _, _ = solve_tsp(op_coords, mode)
            result.append([op_idx[t] for t in tour])
        else:
            result.append(op_idx)
    return result

def generate_gmaps_link(coords, max_wp=10):
    links = []
    for i in range(0, len(coords), max_wp-1):
        seg = coords[i:i+max_wp]
        if len(seg) < 2:
            continue
        pts = "/".join([f"{lat},{lon}" for lat, lon in seg])
        links.append(f"https://www.google.com/maps/dir/{pts}")
    return links

# ==================== FILE FUNCTIONS ====================
def safe_str(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M')
    try:
        return str(val).strip()
    except:
        return ''

def parse_coord(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = safe_str(val).replace(',', '.').replace(' ', '').replace('\xa0', '')
    try:
        return float(s)
    except:
        return None

def read_csv_robust(filepath):
    """Legge CSV con gestione encoding multipla"""
    log_debug(f"Lettura CSV: {filepath}")
    
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']
    content = None
    
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
            log_debug(f"Encoding OK: {enc}")
            break
        except:
            continue
    
    if not content:
        return None, "Errore lettura file (encoding)"
    
    lines = [l for l in content.strip().split('\n') if l.strip()]
    if len(lines) < 2:
        return None, "File vuoto o senza dati"
    
    # Rileva separatore
    first = lines[0]
    sep = ';' if first.count(';') > first.count(',') else ','
    if first.count('\t') > first.count(sep):
        sep = '\t'
    
    # Parse header
    headers = [h.strip().strip('"\'') for h in first.split(sep)]
    
    # Parse righe
    rows = []
    for line in lines[1:]:
        vals = []
        in_quote = False
        current = ''
        for ch in line:
            if ch in '"\'':
                in_quote = not in_quote
            elif ch == sep and not in_quote:
                vals.append(current.strip().strip('"\''))
                current = ''
            else:
                current += ch
        vals.append(current.strip().strip('"\''))
        
        while len(vals) < len(headers):
            vals.append('')
        rows.append(vals[:len(headers)])
    
    log_debug(f"CSV OK: {len(headers)} colonne, {len(rows)} righe")
    return {'headers': headers, 'rows': rows}, headers

def read_xlsx_robust(filepath):
    """Legge XLSX con openpyxl"""
    log_debug(f"Lettura XLSX: {filepath}")
    
    if not OPENPYXL_AVAILABLE:
        return None, f"Libreria openpyxl non disponibile.\n\nUsa file CSV invece di XLSX.\n\nErrore: {OPENPYXL_ERROR}"
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        
        all_rows = []
        for row in ws.iter_rows():
            all_rows.append([safe_str(c.value) for c in row])
        wb.close()
        
        if not all_rows:
            return None, "File vuoto"
        
        headers = [h if h else f"Col{i+1}" for i, h in enumerate(all_rows[0])]
        rows = []
        for row in all_rows[1:]:
            while len(row) < len(headers):
                row.append('')
            rows.append(row[:len(headers)])
        
        log_debug(f"XLSX OK: {len(headers)} colonne, {len(rows)} righe")
        return {'headers': headers, 'rows': rows}, headers
    except Exception as e:
        log_error(f"Errore XLSX: {e}")
        return None, f"Errore lettura Excel: {str(e)}"

def read_file_auto(filepath):
    """Legge file automaticamente in base all'estensione"""
    if not filepath or not os.path.exists(filepath):
        return None, "File non trovato"
    
    if os.path.getsize(filepath) == 0:
        return None, "File vuoto"
    
    if os.path.getsize(filepath) > 50*1024*1024:
        return None, "File troppo grande (max 50MB)"
    
    ext = os.path.splitext(filepath)[1].lower()
    log_debug(f"Estensione file: {ext}")
    
    if ext == '.csv':
        return read_csv_robust(filepath)
    elif ext in ['.xlsx', '.xlsm']:
        if not OPENPYXL_AVAILABLE:
            return None, f"File Excel non supportato!\n\nSalva il file come CSV e riprova.\n\n(openpyxl non disponibile: {OPENPYXL_ERROR})"
        return read_xlsx_robust(filepath)
    elif ext == '.xls':
        return None, "Formato .xls non supportato.\n\nApri il file in Excel e salvalo come CSV."
    else:
        # Prova come CSV
        return read_csv_robust(filepath)

def write_csv(data_rows, filepath):
    """Esporta in CSV"""
    try:
        if not data_rows:
            return False, "Nessun dato"
        
        cols = list(data_rows[0].keys())
        lines = [';'.join(str(c) for c in cols)]
        
        for row in data_rows:
            line = ';'.join(str(row.get(c, '')) for c in cols)
            lines.append(line)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return True, filepath
    except Exception as e:
        return False, str(e)

def export_data(data_rows, filepath):
    """Esporta dati - usa openpyxl se disponibile, altrimenti CSV"""
    if OPENPYXL_AVAILABLE and filepath.endswith('.xlsx'):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if data_rows:
                cols = list(data_rows[0].keys())
                ws.append(cols)
                for row in data_rows:
                    ws.append([row.get(c, '') for c in cols])
            
            wb.save(filepath)
            return True, filepath
        except Exception as e:
            log_error(f"Errore export xlsx: {e}")
            # Fallback a CSV
            csv_path = filepath.replace('.xlsx', '.csv')
            return write_csv(data_rows, csv_path)
    else:
        # Usa CSV
        if filepath.endswith('.xlsx'):
            filepath = filepath.replace('.xlsx', '.csv')
        return write_csv(data_rows, filepath)

# ==================== ANDROID PERMISSIONS ====================
def request_permissions():
    if platform != 'android':
        return
    try:
        from android.permissions import request_permissions, Permission
        request_permissions([
            Permission.READ_EXTERNAL_STORAGE,
            Permission.WRITE_EXTERNAL_STORAGE
        ])
    except Exception as e:
        log_error(f"Permessi: {e}")

def get_storage_paths():
    paths = []
    if platform == 'android':
        try:
            from android.storage import primary_external_storage_path
            p = primary_external_storage_path()
            if p and os.path.exists(p):
                paths.append(p)
                for sub in ['Download', 'Documents']:
                    sp = os.path.join(p, sub)
                    if os.path.exists(sp):
                        paths.append(sp)
        except:
            pass
        for fp in ['/storage/emulated/0', '/storage/emulated/0/Download', '/sdcard', '/sdcard/Download']:
            if os.path.exists(fp) and fp not in paths:
                paths.append(fp)
    else:
        home = os.path.expanduser('~')
        paths.append(home)
        for sub in ['Downloads', 'Documents', 'Desktop']:
            sp = os.path.join(home, sub)
            if os.path.exists(sp):
                paths.append(sp)
    return paths if paths else ['/']

# ==================== MAP WIDGET ====================
class MapWidget(StencilView):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.lat = 41.9
        self.lon = 12.5
        self.zoom = 10
        self.markers = []
        self.route_coords = []
        self.bind(size=self.draw, pos=self.draw)
        Clock.schedule_once(lambda dt: self.draw(), 0.1)
    
    def latlon_to_px(self, lat, lon):
        scale = 256 * (2 ** self.zoom) / 360
        x = (lon - self.lon) * scale + self.width/2 + self.x
        try:
            lat_rad = radians(lat)
            center_rad = radians(self.lat)
            y_scale = 256 * (2 ** self.zoom) / (2 * pi)
            y = (log(tan(pi/4 + center_rad/2)) - log(tan(pi/4 + lat_rad/2))) * y_scale + self.height/2 + self.y
        except:
            y = self.height/2 + self.y
        return x, y
    
    def draw(self, *args):
        self.canvas.clear()
        with self.canvas:
            Color(0.85, 0.9, 0.85, 1)
            Rectangle(pos=self.pos, size=self.size)
            
            Color(0.7, 0.75, 0.7, 0.5)
            sp = max(50, self.width // 8)
            for x in range(int(self.x), int(self.x + self.width), sp):
                Line(points=[x, self.y, x, self.y + self.height], width=1)
            for y in range(int(self.y), int(self.y + self.height), sp):
                Line(points=[self.x, y, self.x + self.width, y], width=1)
            
            if len(self.route_coords) > 1:
                Color(0.2, 0.5, 0.8, 0.8)
                pts = []
                for c in self.route_coords:
                    px, py = self.latlon_to_px(c[0], c[1])
                    pts.extend([px, py])
                if len(pts) >= 4:
                    Line(points=pts, width=2)
            
            for i, m in enumerate(self.markers):
                lat, lon = m[0], m[1]
                px, py = self.latlon_to_px(lat, lon)
                if i == 0:
                    Color(0.2, 0.8, 0.2, 1)
                elif i == len(self.markers) - 1:
                    Color(0.8, 0.2, 0.2, 1)
                else:
                    Color(1, 0.5, 0.2, 1)
                Ellipse(pos=(px-8, py-8), size=(16, 16))
    
    def set_view(self, coords):
        if not coords:
            return
        lats = [c[0] for c in coords]
        lons = [c[1] for c in coords]
        self.lat = (min(lats) + max(lats)) / 2
        self.lon = (min(lons) + max(lons)) / 2
        span = max(max(lats)-min(lats), max(lons)-min(lons))
        self.zoom = 14 if span < 0.01 else 12 if span < 0.1 else 9 if span < 1 else 7 if span < 5 else 5
        self.draw()
    
    def zoom_in(self):
        if self.zoom < 18:
            self.zoom += 1
            self.draw()
    
    def zoom_out(self):
        if self.zoom > 3:
            self.zoom -= 1
            self.draw()

class CardBox(BoxLayout):
    pass

# ==================== KV LAYOUT ====================
KV = '''
<CardBox>:
    orientation: 'vertical'
    padding: 12
    spacing: 8
    canvas.before:
        Color:
            rgba: 0.18, 0.18, 0.22, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [12]

ScreenManager:
    HomeScreen:
    LoadExcelScreen:
    ManualInputScreen:
    ColumnSelectScreen:
    SettingsScreen:
    ProcessingScreen:
    SavedRoutesScreen:
    RouteDetailScreen:
    MapViewScreen:
    ExportScreen:

<HomeScreen>:
    name: 'home'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 70
            padding: 15
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Label:
                text: 'PERCORSI PRO v3.2'
                font_size: '24sp'
                bold: True
                color: 1, 0.5, 0.2, 1
        
        ScrollView:
            BoxLayout:
                orientation: 'vertical'
                padding: 15
                spacing: 15
                size_hint_y: None
                height: self.minimum_height
                
                CardBox:
                    size_hint_y: None
                    height: 180
                    Label:
                        text: 'CARICA DATI'
                        font_size: '16sp'
                        bold: True
                        color: 1, 0.5, 0.2, 1
                        size_hint_y: None
                        height: 30
                    Button:
                        text: '[b]CARICA FILE CSV/EXCEL[/b]'
                        markup: True
                        size_hint_y: None
                        height: 55
                        background_color: 0.2, 0.6, 0.3, 1
                        background_normal: ''
                        on_release: app.go_load_excel()
                    Button:
                        text: 'INSERIMENTO MANUALE'
                        size_hint_y: None
                        height: 55
                        background_color: 0.3, 0.4, 0.6, 1
                        background_normal: ''
                        on_release: app.root.current = 'manual_input'
                
                CardBox:
                    size_hint_y: None
                    height: 120
                    Label:
                        text: 'PERCORSI SALVATI'
                        font_size: '16sp'
                        bold: True
                        color: 0.5, 0.8, 1, 1
                        size_hint_y: None
                        height: 30
                    Button:
                        text: 'VISUALIZZA PERCORSI'
                        size_hint_y: None
                        height: 55
                        background_color: 0.4, 0.4, 0.5, 1
                        background_normal: ''
                        on_release: app.root.current = 'saved_routes'
                
                CardBox:
                    size_hint_y: None
                    height: 120
                    Label:
                        text: 'IMPOSTAZIONI'
                        font_size: '16sp'
                        bold: True
                        color: 0.8, 0.8, 0.5, 1
                        size_hint_y: None
                        height: 30
                    Button:
                        text: 'CONFIGURA'
                        size_hint_y: None
                        height: 55
                        background_color: 0.5, 0.4, 0.3, 1
                        background_normal: ''
                        on_release: app.root.current = 'settings'
                
                Label:
                    text: 'Mattia Prosperi - 2024'
                    font_size: '12sp'
                    color: 0.5, 0.5, 0.5, 1
                    size_hint_y: None
                    height: 30

<LoadExcelScreen>:
    name: 'load_excel'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'home'
            Label:
                text: 'Carica File'
                font_size: '18sp'
                bold: True
        
        # Info formato
        Label:
            id: format_info
            text: 'Formati: CSV (consigliato), XLSX'
            size_hint_y: None
            height: 30
            color: 0.6, 0.8, 0.6, 1
            font_size: '13sp'
        
        BoxLayout:
            orientation: 'vertical'
            padding: 10
            spacing: 10
            
            Spinner:
                id: path_spinner
                text: 'Seleziona cartella...'
                values: []
                size_hint_y: None
                height: 45
                on_text: app.change_path(self.text)
            
            FileChooserListView:
                id: file_chooser
                path: app.get_default_path()
                filters: ['*.csv', '*.CSV', '*.xlsx', '*.XLSX']
                size_hint_y: 0.55
                on_selection: app.on_file_select(self.selection)
            
            Label:
                id: file_status
                text: 'Nessun file selezionato'
                size_hint_y: None
                height: 35
                color: 0.7, 0.7, 0.7, 1
            
            BoxLayout:
                size_hint_y: None
                height: 55
                spacing: 10
                Button:
                    text: 'FILE PICKER'
                    background_color: 0.3, 0.5, 0.6, 1
                    background_normal: ''
                    on_release: app.open_native_picker()
                Button:
                    text: '[b]CARICA[/b]'
                    markup: True
                    background_color: 0.2, 0.7, 0.3, 1
                    background_normal: ''
                    on_release: app.load_selected_file()

<ManualInputScreen>:
    name: 'manual_input'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'home'
            Label:
                text: 'Input Manuale'
                font_size: '18sp'
                bold: True
        
        ScrollView:
            BoxLayout:
                orientation: 'vertical'
                padding: 15
                spacing: 10
                size_hint_y: None
                height: self.minimum_height
                
                Label:
                    text: 'Coordinate (lat,lon per riga):'
                    size_hint_y: None
                    height: 25
                
                TextInput:
                    id: coords_input
                    hint_text: '45.4642,9.1900\\n41.9028,12.4964'
                    multiline: True
                    size_hint_y: None
                    height: 180
                    background_color: 0.15, 0.15, 0.18, 1
                    foreground_color: 1, 1, 1, 1
                
                Label:
                    text: 'Etichette (opzionale):'
                    size_hint_y: None
                    height: 25
                
                TextInput:
                    id: labels_input
                    hint_text: 'Milano\\nRoma'
                    multiline: True
                    size_hint_y: None
                    height: 100
                    background_color: 0.15, 0.15, 0.18, 1
                    foreground_color: 1, 1, 1, 1
                
                BoxLayout:
                    size_hint_y: None
                    height: 45
                    spacing: 10
                    Button:
                        text: 'ESEMPIO'
                        background_color: 0.4, 0.4, 0.5, 1
                        background_normal: ''
                        on_release: app.load_example()
                    Button:
                        text: 'PULISCI'
                        background_color: 0.5, 0.3, 0.3, 1
                        background_normal: ''
                        on_release: app.clear_input()
                
                Label:
                    id: coord_count
                    text: 'Coordinate: 0'
                    size_hint_y: None
                    height: 30
                
                Button:
                    text: '[b]OTTIMIZZA PERCORSO[/b]'
                    markup: True
                    size_hint_y: None
                    height: 55
                    background_color: 1, 0.2, 0.4, 1
                    background_normal: ''
                    on_release: app.start_manual_opt()

<ColumnSelectScreen>:
    name: 'column_select'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'load_excel'
            Label:
                text: 'Seleziona Colonne'
                font_size: '18sp'
                bold: True
        
        ScrollView:
            BoxLayout:
                orientation: 'vertical'
                padding: 15
                spacing: 15
                size_hint_y: None
                height: self.minimum_height
                
                Label:
                    id: loaded_info
                    text: 'File caricato'
                    size_hint_y: None
                    height: 30
                    color: 0.5, 0.8, 0.5, 1
                
                CardBox:
                    size_hint_y: None
                    height: 160
                    Label:
                        text: 'COLONNE OBBLIGATORIE'
                        bold: True
                        color: 1, 0.5, 0.2, 1
                        size_hint_y: None
                        height: 30
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Latitudine:'
                            size_hint_x: 0.4
                        Spinner:
                            id: lat_col
                            text: 'Seleziona...'
                            size_hint_x: 0.6
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Longitudine:'
                            size_hint_x: 0.4
                        Spinner:
                            id: lon_col
                            text: 'Seleziona...'
                            size_hint_x: 0.6
                
                CardBox:
                    size_hint_y: None
                    height: 220
                    Label:
                        text: 'COLONNE OPZIONALI'
                        bold: True
                        color: 0.5, 0.8, 1, 1
                        size_hint_y: None
                        height: 30
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Etichetta:'
                            size_hint_x: 0.4
                        Spinner:
                            id: label_col
                            text: '-- Nessuna --'
                            size_hint_x: 0.6
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Indirizzo:'
                            size_hint_x: 0.4
                        Spinner:
                            id: addr_col
                            text: '-- Nessuna --'
                            size_hint_x: 0.6
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Operatore:'
                            size_hint_x: 0.4
                        Spinner:
                            id: op_col
                            text: '-- Nessuna --'
                            size_hint_x: 0.6
                
                Button:
                    text: '[b]AVVIA OTTIMIZZAZIONE[/b]'
                    markup: True
                    size_hint_y: None
                    height: 60
                    background_color: 1, 0.2, 0.4, 1
                    background_normal: ''
                    on_release: app.start_excel_opt()

<SettingsScreen>:
    name: 'settings'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'home'
            Label:
                text: 'Impostazioni'
                font_size: '18sp'
                bold: True
        
        ScrollView:
            BoxLayout:
                orientation: 'vertical'
                padding: 15
                spacing: 15
                size_hint_y: None
                height: self.minimum_height
                
                CardBox:
                    size_hint_y: None
                    height: 120
                    Label:
                        text: 'CALCOLO DISTANZE'
                        bold: True
                        color: 1, 0.5, 0.2, 1
                        size_hint_y: None
                        height: 30
                    Spinner:
                        id: dist_mode
                        text: 'Haversine (veloce)'
                        values: ['Haversine (veloce)', 'OSRM (strade reali)']
                        size_hint_y: None
                        height: 45
                
                CardBox:
                    size_hint_y: None
                    height: 200
                    Label:
                        text: 'MULTI-OPERATORE'
                        bold: True
                        color: 0.5, 0.8, 1, 1
                        size_hint_y: None
                        height: 30
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        CheckBox:
                            id: multi_op
                            active: False
                            size_hint_x: 0.2
                        Label:
                            text: 'Dividi per operatori'
                            halign: 'left'
                            text_size: self.size
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Num. operatori:'
                            size_hint_x: 0.6
                        TextInput:
                            id: num_ops
                            text: '2'
                            input_filter: 'int'
                            multiline: False
                            size_hint_x: 0.4
                            background_color: 0.2, 0.2, 0.25, 1
                            foreground_color: 1, 1, 1, 1
                    BoxLayout:
                        size_hint_y: None
                        height: 40
                        Label:
                            text: 'Tappe/operatore:'
                            size_hint_x: 0.6
                        TextInput:
                            id: items_op
                            text: '15'
                            input_filter: 'int'
                            multiline: False
                            size_hint_x: 0.4
                            background_color: 0.2, 0.2, 0.25, 1
                            foreground_color: 1, 1, 1, 1
                
                CardBox:
                    size_hint_y: None
                    height: 140
                    Label:
                        text: 'STATO SISTEMA'
                        bold: True
                        color: 0.8, 0.8, 0.5, 1
                        size_hint_y: None
                        height: 30
                    Label:
                        id: sys_status
                        text: 'Verifica...'
                        font_size: '12sp'
                        size_hint_y: None
                        height: 80
                        halign: 'left'
                        text_size: self.size

<ProcessingScreen>:
    name: 'processing'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        Widget:
            size_hint_y: 0.3
        
        BoxLayout:
            orientation: 'vertical'
            padding: 30
            spacing: 20
            size_hint_y: 0.4
            
            Label:
                text: 'OTTIMIZZAZIONE'
                font_size: '20sp'
                bold: True
                color: 1, 0.5, 0.2, 1
            
            ProgressBar:
                id: progress
                max: 100
                value: 0
            
            Label:
                id: progress_pct
                text: '0%'
                font_size: '24sp'
                bold: True
            
            Label:
                id: progress_txt
                text: 'Preparazione...'
                color: 0.7, 0.7, 0.7, 1
        
        Widget:
            size_hint_y: 0.3

<SavedRoutesScreen>:
    name: 'saved_routes'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'home'
            Label:
                text: 'Percorsi Salvati'
                font_size: '18sp'
                bold: True
        
        ScrollView:
            BoxLayout:
                id: routes_list
                orientation: 'vertical'
                padding: 10
                spacing: 8
                size_hint_y: None
                height: self.minimum_height

<RouteDetailScreen>:
    name: 'route_detail'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'saved_routes'
            Label:
                id: route_title
                text: 'Dettaglio'
                font_size: '16sp'
                bold: True
        
        Label:
            id: route_info
            text: 'Distanza: -- | Tappe: --'
            size_hint_y: None
            height: 35
            color: 0.7, 0.7, 0.7, 1
        
        ScrollView:
            BoxLayout:
                id: stops_list
                orientation: 'vertical'
                padding: 10
                spacing: 5
                size_hint_y: None
                height: self.minimum_height
        
        BoxLayout:
            size_hint_y: None
            height: 55
            padding: 8
            spacing: 8
            Button:
                text: 'MAPPA'
                background_color: 0.3, 0.5, 0.7, 1
                background_normal: ''
                on_release: app.show_map()
            Button:
                text: 'EXPORT'
                background_color: 0.2, 0.6, 0.3, 1
                background_normal: ''
                on_release: app.root.current = 'export'
            Button:
                text: 'NAVIGA'
                background_color: 0.7, 0.3, 0.3, 1
                background_normal: ''
                on_release: app.start_nav()

<MapViewScreen>:
    name: 'map_view'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'route_detail'
            Label:
                text: 'Mappa'
                font_size: '18sp'
                bold: True
        
        MapWidget:
            id: map_widget
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 5
            spacing: 10
            Button:
                text: '-'
                size_hint_x: 0.25
                background_color: 0.3, 0.3, 0.4, 1
                background_normal: ''
                on_release: map_widget.zoom_out()
            Button:
                text: 'CENTRA'
                size_hint_x: 0.5
                background_color: 0.3, 0.5, 0.3, 1
                background_normal: ''
                on_release: app.center_map()
            Button:
                text: '+'
                size_hint_x: 0.25
                background_color: 0.3, 0.3, 0.4, 1
                background_normal: ''
                on_release: map_widget.zoom_in()

<ExportScreen>:
    name: 'export'
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.12, 1
            Rectangle:
                pos: self.pos
                size: self.size
        
        BoxLayout:
            size_hint_y: None
            height: 50
            padding: 10
            canvas.before:
                Color:
                    rgba: 0.15, 0.15, 0.18, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: '< INDIETRO'
                size_hint_x: 0.3
                background_color: 0.3, 0.3, 0.35, 1
                background_normal: ''
                on_release: app.root.current = 'route_detail'
            Label:
                text: 'Esporta'
                font_size: '18sp'
                bold: True
        
        ScrollView:
            BoxLayout:
                orientation: 'vertical'
                padding: 15
                spacing: 15
                size_hint_y: None
                height: self.minimum_height
                
                CardBox:
                    size_hint_y: None
                    height: 200
                    Label:
                        text: 'FORMATO'
                        bold: True
                        color: 1, 0.5, 0.2, 1
                        size_hint_y: None
                        height: 30
                    Button:
                        text: 'CSV / EXCEL'
                        size_hint_y: None
                        height: 50
                        background_color: 0.2, 0.6, 0.3, 1
                        background_normal: ''
                        on_release: app.do_export_excel()
                    Button:
                        text: 'GPX (Navigatori)'
                        size_hint_y: None
                        height: 50
                        background_color: 0.3, 0.4, 0.6, 1
                        background_normal: ''
                        on_release: app.do_export_gpx()
                    Button:
                        text: 'KML (Google Earth)'
                        size_hint_y: None
                        height: 50
                        background_color: 0.5, 0.4, 0.3, 1
                        background_normal: ''
                        on_release: app.do_export_kml()
                
                CardBox:
                    size_hint_y: None
                    height: 80
                    Button:
                        text: '[b]APRI GOOGLE MAPS[/b]'
                        markup: True
                        size_hint_y: None
                        height: 55
                        background_color: 0.2, 0.6, 0.9, 1
                        background_normal: ''
                        on_release: app.start_nav()
'''

# ==================== SCREENS ====================
class HomeScreen(Screen): pass
class LoadExcelScreen(Screen): pass
class ManualInputScreen(Screen): pass
class ColumnSelectScreen(Screen): pass
class SettingsScreen(Screen): pass
class ProcessingScreen(Screen): pass
class SavedRoutesScreen(Screen): pass
class RouteDetailScreen(Screen): pass
class MapViewScreen(Screen): pass
class ExportScreen(Screen): pass

# ==================== MAIN APP ====================
class PercorsiApp(App):
    excel_data = None
    excel_cols = []
    coords = ListProperty([])
    labels = ListProperty([])
    orig_data = []
    current_route = None
    saved_routes = ListProperty([])
    selected_file = StringProperty('')
    
    def build(self):
        self.title = 'Percorsi Pro v3.2'
        if platform not in ('android', 'ios'):
            Window.size = (400, 750)
        self.load_routes()
        return Builder.load_string(KV)
    
    def on_start(self):
        if platform == 'android':
            Clock.schedule_once(lambda dt: request_permissions(), 0.5)
        Clock.schedule_once(lambda dt: self._init_ui(), 1)
    
    def _init_ui(self):
        try:
            self.refresh_routes_ui()
            self._update_sys_status()
            self._init_paths()
            self._update_format_info()
        except Exception as e:
            log_error(f"Init: {e}")
    
    def _update_format_info(self):
        """Aggiorna info formati supportati"""
        try:
            scr = self.root.get_screen('load_excel')
            if OPENPYXL_AVAILABLE:
                scr.ids.format_info.text = "Formati: CSV, XLSX"
                scr.ids.format_info.color = (0.5, 0.8, 0.5, 1)
            else:
                scr.ids.format_info.text = "Solo CSV (XLSX non supportato)"
                scr.ids.format_info.color = (1, 0.7, 0.3, 1)
        except:
            pass
    
    def _update_sys_status(self):
        try:
            scr = self.root.get_screen('settings')
            xlsx_status = "OK" if OPENPYXL_AVAILABLE else f"NO ({OPENPYXL_ERROR[:30]})"
            lines = [
                f"Excel (openpyxl): {xlsx_status}",
                f"CSV: OK",
                f"Rete (requests): {'OK' if REQUESTS_AVAILABLE else 'NO'}",
                f"Platform: {platform}"
            ]
            scr.ids.sys_status.text = '\n'.join(lines)
        except:
            pass
    
    def _init_paths(self):
        try:
            scr = self.root.get_screen('load_excel')
            paths = get_storage_paths()
            scr.ids.path_spinner.values = paths
            if paths:
                scr.ids.path_spinner.text = paths[0]
        except:
            pass
    
    def get_default_path(self):
        paths = get_storage_paths()
        return paths[0] if paths else '/'
    
    # ==================== FILE LOADING ====================
    def go_load_excel(self):
        self._init_paths()
        self._update_format_info()
        self.root.current = 'load_excel'
    
    def change_path(self, path):
        try:
            if os.path.exists(path) and os.path.isdir(path):
                scr = self.root.get_screen('load_excel')
                scr.ids.file_chooser.path = path
        except:
            pass
    
    def on_file_select(self, selection):
        try:
            scr = self.root.get_screen('load_excel')
            if selection:
                fp = selection[0]
                name = os.path.basename(fp)
                size = os.path.getsize(fp) / 1024
                ext = os.path.splitext(fp)[1].lower()
                
                # Avvisa se xlsx e openpyxl non disponibile
                if ext == '.xlsx' and not OPENPYXL_AVAILABLE:
                    scr.ids.file_status.text = f"{name} - XLSX non supportato! Usa CSV"
                    scr.ids.file_status.color = (1, 0.4, 0.4, 1)
                else:
                    scr.ids.file_status.text = f"Sel: {name} ({size:.1f} KB)"
                    scr.ids.file_status.color = (0.7, 0.7, 0.7, 1)
                
                self.selected_file = fp
            else:
                scr.ids.file_status.text = "Nessun file selezionato"
                self.selected_file = ''
        except:
            pass
    
    def open_native_picker(self):
        if not PLYER_AVAILABLE:
            self.popup("Info", "File picker nativo non disponibile.\n\nUsa il selettore file sopra.")
            return
        try:
            plyer_fc.open_file(
                on_selection=self._on_native_select,
                filters=[("CSV/Excel", "*.csv", "*.xlsx")]
            )
        except Exception as e:
            self.popup("Errore", str(e))
    
    def _on_native_select(self, selection):
        if selection:
            self.selected_file = selection[0]
            Clock.schedule_once(lambda dt: self._do_load(self.selected_file), 0.1)
    
    def load_selected_file(self):
        try:
            scr = self.root.get_screen('load_excel')
            sel = scr.ids.file_chooser.selection
            if sel:
                self.selected_file = sel[0]
            
            if not self.selected_file:
                self.popup("Errore", "Seleziona un file!")
                return
            
            self._do_load(self.selected_file)
        except Exception as e:
            self.popup("Errore", str(e))
    
    def _do_load(self, filepath):
        if not filepath or not os.path.exists(filepath):
            self.popup("Errore", "File non trovato")
            return
        
        ext = os.path.splitext(filepath)[1].lower()
        
        # Check xlsx senza openpyxl
        if ext == '.xlsx' and not OPENPYXL_AVAILABLE:
            self.popup("Formato non supportato", 
                      "I file XLSX non sono supportati su questo dispositivo.\n\n"
                      "SOLUZIONE:\n"
                      "1. Apri il file in Excel\n"
                      "2. Salva come CSV\n"
                      "3. Carica il file CSV")
            return
        
        if ext not in ['.xlsx', '.csv', '.xlsm']:
            self.popup("Errore", f"Formato non supportato: {ext}\n\nUsa CSV o XLSX")
            return
        
        self._show_loading()
        
        def do():
            try:
                data, cols = read_file_auto(filepath)
                Clock.schedule_once(lambda dt: self._on_loaded(data, cols, filepath), 0)
            except Exception as e:
                log_error(f"Load error: {e}")
                Clock.schedule_once(lambda dt: self._on_load_error(str(e)), 0)
        
        threading.Thread(target=do, daemon=True).start()
    
    def _show_loading(self):
        content = BoxLayout(orientation='vertical', padding=10)
        content.add_widget(Label(text='Caricamento...'))
        content.add_widget(ProgressBar(max=100, value=50))
        self._loading = Popup(title='Attendere', content=content, size_hint=(0.8, 0.3), auto_dismiss=False)
        self._loading.open()
    
    def _hide_loading(self):
        try:
            if hasattr(self, '_loading'):
                self._loading.dismiss()
        except:
            pass
    
    def _on_loaded(self, data, cols, filepath):
        self._hide_loading()
        if data is None:
            self.popup("Errore Caricamento", str(cols))
            return
        
        self.excel_data = data
        self.excel_cols = cols
        self.root.current = 'column_select'
        
        try:
            scr = self.root.get_screen('column_select')
            name = os.path.basename(filepath)
            nrows = len(data.get('rows', []))
            scr.ids.loaded_info.text = f"{name} ({nrows} righe, {len(cols)} col)"
        except:
            pass
        
        Clock.schedule_once(lambda dt: self._setup_spinners(cols), 0.2)
    
    def _on_load_error(self, msg):
        self._hide_loading()
        self.popup("Errore", msg)
    
    def _setup_spinners(self, cols):
        try:
            scr = self.root.get_screen('column_select')
            opts = ['-- Nessuna --'] + cols
            
            scr.ids.lat_col.values = cols
            scr.ids.lon_col.values = cols
            scr.ids.label_col.values = opts
            scr.ids.addr_col.values = opts
            scr.ids.op_col.values = opts
            
            # Auto-detect
            for c in cols:
                u = c.upper()
                if 'LAT' in u:
                    scr.ids.lat_col.text = c
                elif 'LON' in u:
                    scr.ids.lon_col.text = c
                elif any(x in u for x in ['NOME', 'NAME', 'LABEL', 'CLIENTE']):
                    scr.ids.label_col.text = c
                elif any(x in u for x in ['INDIRIZZO', 'ADDRESS', 'VIA']):
                    scr.ids.addr_col.text = c
                elif 'OPERATOR' in u or 'OPERATORE' in u:
                    scr.ids.op_col.text = c
        except Exception as e:
            log_error(f"Spinners: {e}")
    
    # ==================== OPTIMIZATION ====================
    def start_excel_opt(self):
        scr = self.root.get_screen('column_select')
        lat_c = scr.ids.lat_col.text
        lon_c = scr.ids.lon_col.text
        
        if lat_c == 'Seleziona...' or lon_c == 'Seleziona...':
            self.popup("Errore", "Seleziona Latitudine e Longitudine!")
            return
        
        self.coords = []
        self.labels = []
        self.orig_data = []
        
        label_c = scr.ids.label_col.text
        addr_c = scr.ids.addr_col.text
        
        if self.excel_data and 'headers' in self.excel_data:
            hdrs = self.excel_data['headers']
            rows = self.excel_data['rows']
            
            try:
                lat_i = hdrs.index(lat_c)
                lon_i = hdrs.index(lon_c)
            except:
                self.popup("Errore", "Colonne non trovate")
                return
            
            label_i = hdrs.index(label_c) if label_c in hdrs else -1
            addr_i = hdrs.index(addr_c) if addr_c in hdrs else -1
            
            for row in rows:
                lat = parse_coord(row[lat_i] if lat_i < len(row) else None)
                lon = parse_coord(row[lon_i] if lon_i < len(row) else None)
                
                if lat is None or lon is None:
                    continue
                if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                    continue
                
                self.coords.append((lat, lon))
                
                rd = {h: safe_str(row[i] if i < len(row) else '') for i, h in enumerate(hdrs)}
                self.orig_data.append(rd)
                
                parts = []
                if label_i >= 0 and label_i < len(row) and row[label_i]:
                    parts.append(safe_str(row[label_i]))
                if addr_i >= 0 and addr_i < len(row) and row[addr_i]:
                    parts.append(safe_str(row[addr_i]))
                self.labels.append(' - '.join(parts) if parts else f"Punto {len(self.coords)}")
        
        if len(self.coords) < 2:
            self.popup("Errore", f"Solo {len(self.coords)} coordinate valide.\n\nVerifica le colonne selezionate.")
            return
        
        self._run_opt()
    
    def load_example(self):
        scr = self.root.get_screen('manual_input')
        scr.ids.coords_input.text = "45.4642,9.1900\n41.9028,12.4964\n43.7696,11.2558\n44.4949,11.3426\n40.8518,14.2681"
        scr.ids.labels_input.text = "Milano\nRoma\nFirenze\nBologna\nNapoli"
        scr.ids.coord_count.text = "Coordinate: 5"
    
    def clear_input(self):
        scr = self.root.get_screen('manual_input')
        scr.ids.coords_input.text = ""
        scr.ids.labels_input.text = ""
        scr.ids.coord_count.text = "Coordinate: 0"
    
    def start_manual_opt(self):
        scr = self.root.get_screen('manual_input')
        text = scr.ids.coords_input.text
        
        self.coords = []
        for line in text.strip().split('\n'):
            line = line.strip()
            if not line:
                continue
            for sep in [',', ';', '\t', ' ']:
                if sep in line:
                    parts = line.split(sep)
                    if len(parts) >= 2:
                        lat = parse_coord(parts[0])
                        lon = parse_coord(parts[1])
                        if lat and lon and -90 <= lat <= 90 and -180 <= lon <= 180:
                            self.coords.append((lat, lon))
                            break
        
        labels_txt = scr.ids.labels_input.text
        self.labels = [l.strip() for l in labels_txt.split('\n') if l.strip()]
        while len(self.labels) < len(self.coords):
            self.labels.append(f"Punto {len(self.labels)+1}")
        
        self.orig_data = []
        
        if len(self.coords) < 2:
            self.popup("Errore", "Minimo 2 coordinate!")
            return
        
        self._run_opt()
    
    def _run_opt(self):
        self.root.current = 'processing'
        Clock.schedule_once(lambda dt: self._reset_progress(), 0.1)
        threading.Thread(target=self._do_opt, daemon=True).start()
    
    def _reset_progress(self):
        try:
            scr = self.root.get_screen('processing')
            scr.ids.progress.value = 0
            scr.ids.progress_pct.text = "0%"
            scr.ids.progress_txt.text = "Preparazione..."
        except:
            pass
    
    def _set_progress(self, p):
        try:
            scr = self.root.get_screen('processing')
            scr.ids.progress.value = int(p * 100)
            scr.ids.progress_pct.text = f"{int(p*100)}%"
        except:
            pass
    
    def _do_opt(self):
        def cb(p):
            Clock.schedule_once(lambda dt: self._set_progress(p), 0)
        
        try:
            mode = self._get_mode()
            multi, num_ops, items_per = self._get_multi_settings()
            
            all_idx = list(range(len(self.coords)))
            
            if multi and num_ops > 1:
                groups = divide_for_operators(all_idx, list(self.coords), num_ops, items_per, mode)
                routes = []
                for oi, op_idx in enumerate(groups):
                    if not op_idx:
                        continue
                    oc = [self.coords[i] for i in op_idx]
                    ol = [self.labels[i] for i in op_idx]
                    od = [self.orig_data[i] if i < len(self.orig_data) else {} for i in op_idx]
                    
                    if len(oc) > 1:
                        tour, total, dists = solve_tsp(oc, mode, cb)
                        oc = [oc[t] for t in tour]
                        ol = [ol[t] for t in tour]
                        od = [od[t] for t in tour]
                    else:
                        total = 0
                        dists = [0]
                    
                    routes.append({
                        'id': f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{oi+1}",
                        'name': f"Op{oi+1} - {datetime.now().strftime('%d/%m %H:%M')}",
                        'operator': f"Operatore {oi+1}",
                        'coords': oc, 'labels': ol, 'orig_data': od,
                        'distances': dists, 'total': total, 'stops': len(oc)
                    })
                
                Clock.schedule_once(lambda dt: self._opt_done_multi(routes), 0)
            else:
                tour, total, dists = solve_tsp(list(self.coords), mode, cb)
                oc = [self.coords[t] for t in tour]
                ol = [self.labels[t] for t in tour]
                od = [self.orig_data[t] if t < len(self.orig_data) else {} for t in tour]
                
                route = {
                    'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
                    'name': f"Percorso {datetime.now().strftime('%d/%m %H:%M')}",
                    'coords': oc, 'labels': ol, 'orig_data': od,
                    'distances': dists, 'total': total, 'stops': len(oc)
                }
                Clock.schedule_once(lambda dt: self._opt_done(route), 0)
        except Exception as e:
            log_error(f"Opt: {e}")
            Clock.schedule_once(lambda dt: self._opt_error(str(e)), 0)
    
    def _get_mode(self):
        try:
            scr = self.root.get_screen('settings')
            txt = scr.ids.dist_mode.text
            return "osrm" if "OSRM" in txt else "haversine"
        except:
            return "haversine"
    
    def _get_multi_settings(self):
        try:
            scr = self.root.get_screen('settings')
            enabled = scr.ids.multi_op.active
            num = int(scr.ids.num_ops.text or '2')
            items = int(scr.ids.items_op.text or '15')
            return enabled, max(1, num), max(1, items)
        except:
            return False, 2, 15
    
    def _opt_done(self, route):
        self.current_route = route
        self.saved_routes.append(route)
        self.save_routes()
        self.refresh_routes_ui()
        self.show_route(route)
    
    def _opt_done_multi(self, routes):
        for r in routes:
            self.saved_routes.append(r)
        self.save_routes()
        self.refresh_routes_ui()
        if routes:
            self.current_route = routes[0]
            self.show_route(routes[0])
        else:
            self.root.current = 'saved_routes'
    
    def _opt_error(self, msg):
        self.popup("Errore", msg)
        self.root.current = 'home'
    
    # ==================== ROUTES ====================
    def show_route(self, route):
        self.current_route = route
        scr = self.root.get_screen('route_detail')
        scr.ids.route_title.text = route['name']
        scr.ids.route_info.text = f"Distanza: {self._fmt_dist(route['total'])} | Tappe: {route['stops']}"
        
        sl = scr.ids.stops_list
        sl.clear_widgets()
        
        cum = 0
        for i, (c, l) in enumerate(zip(route['coords'], route['labels'])):
            d = route['distances'][i] if i < len(route['distances']) else 0
            cum += d
            
            box = BoxLayout(size_hint_y=None, height=45, padding=5)
            box.add_widget(Label(text=str(i+1), size_hint_x=0.1, bold=True, color=(1,0.5,0.2,1)))
            box.add_widget(Label(text=l[:35], size_hint_x=0.6, halign='left', text_size=(200,None)))
            box.add_widget(Label(text=self._fmt_dist(cum), size_hint_x=0.3, color=(0.5,0.8,0.5,1)))
            sl.add_widget(box)
        
        self.root.current = 'route_detail'
        self.refresh_routes_ui()
    
    def _fmt_dist(self, m):
        return f"{m/1000:.1f} km" if m >= 1000 else f"{int(m)} m"
    
    def refresh_routes_ui(self):
        try:
            scr = self.root.get_screen('saved_routes')
            rl = scr.ids.routes_list
            rl.clear_widgets()
            
            if not self.saved_routes:
                rl.add_widget(Label(text='Nessun percorso salvato', size_hint_y=None, height=50))
                return
            
            for r in reversed(self.saved_routes[-20:]):
                box = BoxLayout(size_hint_y=None, height=70, padding=5, spacing=5)
                
                info = BoxLayout(orientation='vertical', size_hint_x=0.7)
                info.add_widget(Label(text=r['name'], bold=True, halign='left', text_size=(250,None)))
                info.add_widget(Label(text=f"{r['stops']} tappe - {self._fmt_dist(r['total'])}", 
                                     font_size='12sp', color=(0.6,0.6,0.6,1)))
                
                btn = Button(text='APRI', size_hint_x=0.3, background_color=(0.3,0.5,0.6,1), background_normal='')
                btn.bind(on_release=lambda x, rt=r: self.show_route(rt))
                
                box.add_widget(info)
                box.add_widget(btn)
                rl.add_widget(box)
        except Exception as e:
            log_error(f"RefreshUI: {e}")
    
    # ==================== MAP ====================
    def show_map(self):
        if not self.current_route:
            return
        scr = self.root.get_screen('map_view')
        mw = scr.ids.map_widget
        mw.markers = [(c[0], c[1], l) for c, l in zip(self.current_route['coords'], self.current_route['labels'])]
        mw.route_coords = self.current_route['coords']
        mw.set_view(self.current_route['coords'])
        self.root.current = 'map_view'
    
    def center_map(self):
        if self.current_route:
            scr = self.root.get_screen('map_view')
            scr.ids.map_widget.set_view(self.current_route['coords'])
    
    # ==================== NAVIGATION ====================
    def start_nav(self):
        if not self.current_route:
            return
        links = generate_gmaps_link(self.current_route['coords'])
        if links:
            import webbrowser
            webbrowser.open(links[0])
    
    # ==================== EXPORT ====================
    def _get_export_path(self, fname):
        if platform == 'android':
            try:
                from android.storage import primary_external_storage_path
                return os.path.join(primary_external_storage_path(), 'Download', fname)
            except:
                pass
        return os.path.join(os.path.expanduser('~'), fname)
    
    def do_export_excel(self):
        if not self.current_route:
            return
        
        rows = []
        cum = 0
        for i, (c, l) in enumerate(zip(self.current_route['coords'], self.current_route['labels'])):
            d = self.current_route['distances'][i] if i < len(self.current_route['distances']) else 0
            cum += d
            row = {'N': i+1, 'Lat': c[0], 'Lon': c[1], 'Label': l, 'Dist_m': d, 'Cum_m': cum}
            od = self.current_route.get('orig_data', [])
            if i < len(od) and od[i]:
                for k, v in od[i].items():
                    if k not in row:
                        row[k] = v
            rows.append(row)
        
        # Usa xlsx se disponibile, altrimenti csv
        ext = '.xlsx' if OPENPYXL_AVAILABLE else '.csv'
        fp = self._get_export_path(f"percorso_{self.current_route['id']}{ext}")
        ok, res = export_data(rows, fp)
        self.popup("Export" if ok else "Errore", res)
    
    def do_export_gpx(self):
        if not self.current_route:
            return
        
        gpx = ['<?xml version="1.0" encoding="UTF-8"?>', '<gpx version="1.1">']
        for c, l in zip(self.current_route['coords'], self.current_route['labels']):
            gpx.append(f'<wpt lat="{c[0]}" lon="{c[1]}"><name>{l}</name></wpt>')
        gpx.append('<trk><trkseg>')
        for c in self.current_route['coords']:
            gpx.append(f'<trkpt lat="{c[0]}" lon="{c[1]}"/>')
        gpx.append('</trkseg></trk></gpx>')
        
        fp = self._get_export_path(f"percorso_{self.current_route['id']}.gpx")
        try:
            with open(fp, 'w', encoding='utf-8') as f:
                f.write('\n'.join(gpx))
            self.popup("Export", fp)
        except Exception as e:
            self.popup("Errore", str(e))
    
    def do_export_kml(self):
        if not self.current_route:
            return
        
        kml = ['<?xml version="1.0" encoding="UTF-8"?>', '<kml xmlns="http://www.opengis.net/kml/2.2">', '<Document>']
        for c, l in zip(self.current_route['coords'], self.current_route['labels']):
            kml.append(f'<Placemark><name>{l}</name><Point><coordinates>{c[1]},{c[0]},0</coordinates></Point></Placemark>')
        coords_str = ' '.join([f"{c[1]},{c[0]},0" for c in self.current_route['coords']])
        kml.append(f'<Placemark><name>Percorso</name><LineString><coordinates>{coords_str}</coordinates></LineString></Placemark>')
        kml.append('</Document></kml>')
        
        fp = self._get_export_path(f"percorso_{self.current_route['id']}.kml")
        try:
            with open(fp, 'w', encoding='utf-8') as f:
                f.write('\n'.join(kml))
            self.popup("Export", fp)
        except Exception as e:
            self.popup("Errore", str(e))
    
    # ==================== PERSISTENCE ====================
    def _get_data_path(self):
        if platform == 'android':
            return os.path.join(self.user_data_dir, 'routes.json')
        return os.path.join(os.path.expanduser('~'), '.percorsi_routes.json')
    
    def save_routes(self):
        try:
            data = []
            for r in self.saved_routes[-50:]:
                data.append({k: r[k] for k in ['id','name','coords','labels','distances','total','stops'] if k in r})
            with open(self._get_data_path(), 'w') as f:
                json.dump(data, f)
        except Exception as e:
            log_error(f"Save: {e}")
    
    def load_routes(self):
        try:
            fp = self._get_data_path()
            if os.path.exists(fp):
                with open(fp, 'r') as f:
                    self.saved_routes = json.load(f)
        except:
            self.saved_routes = []
    
    # ==================== UTILS ====================
    def popup(self, title, msg):
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        lbl = Label(text=str(msg), text_size=(300, None), halign='center')
        lbl.bind(texture_size=lbl.setter('size'))
        content.add_widget(lbl)
        btn = Button(text='OK', size_hint_y=None, height=45, background_color=(0.3,0.5,0.7,1), background_normal='')
        content.add_widget(btn)
        p = Popup(title=title, content=content, size_hint=(0.85, 0.5))
        btn.bind(on_release=p.dismiss)
        p.open()


if __name__ == '__main__':
    PercorsiApp().run()
